import openpyxl

if __name__ == '__main__':
    # Initialize Workbook
    workbook = openpyxl.Workbook()
    # Or load old workbook with
    # workbook = openpyxl.load_workbook("hello_world.xlsx")
    sheet = workbook.active

    # Access rows and collumns list_of_numbers
    MAX_ROWS = sheet.rows
    MAX_COLUMNS = sheet.columns

    # Access cell as name
    sheet["A1"] = "hello"
    sheet["B1"] = "world!"

    # Access cell as row and collunm
    sheet.cell(row=2, column=1).value = "hello"
    print(sheet.cell(row=2, column=1).value)
    # Save workbook
    workbook.save(filename="hello_world.xlsx")